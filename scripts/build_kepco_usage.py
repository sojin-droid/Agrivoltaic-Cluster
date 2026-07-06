#!/usr/bin/env python3
"""metadata/kepco_usage/*.xlsx(한전 빅데이터센터 "산업분류별 전력사용량", 수동 로그인
다운로드, 2023년 01월~2025년 12월 3개년 누적)를 읽어 시군 카드 코드 단위 연평균
GWh로 환산한 kepco_industrial_usage.json(루트, 커밋 대상, 경량)을 만든다.

원본 xlsx는 로그인 세션이 있어야 받을 수 있어 스크립트로 재생성 불가 —
metadata/kepco_usage/는 gitignore, 이 스크립트의 산출물만 커밋한다.

파일명의 첫 단어를 시군명으로 파싱해 CODE_MAP으로 시군 카드 코드에 매핑한다.
시흥시·안산시는 카드가 하나(41390)라 시흥시 파일 + 안산시 (1)/(2) 파일 3개를 합산한다
(안산 (1)/(2)는 총사용량이 서로 달라 동일 데이터의 표기 차이가 아니라 서로 다른
하위 지역 집계로 판단 — 상록구/단원구로 추정, 확인 안 됨. 합산해서 안산시 전체로 취급).
용인시도 동일 사유로 (1)/(2) 두 파일을 합산한다(2026-07-06 재다운로드로 확인 —
최초 받은 파일은 헤더가 "시도: 전체(시도), 시군구: 전체(시/군/구)"인 전국 집계
오export였어서 재요청, BROKEN_FILES는 이제 비어있지만 향후 유사 사고 재발 시
같은 방식으로 등록해 두면 스크립트가 자동으로 해당 시군만 제외한다).
"""
import glob
import json
import os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
USAGE_DIR = os.path.join(ROOT, "metadata", "kepco_usage")
OUT_PATH = os.path.join(ROOT, "kepco_industrial_usage.json")

PERIOD_LABEL = "2023.01–2025.12(3개년 누적치의 연평균)"
SOURCE_LABEL = "한국전력공사 빅데이터센터(bigdata.kepco.co.kr), 산업분류별 전력사용량 2023–2025 연평균"
CAVEAT = ("시군구 전체 업종(KSIC) 합산치로, 본 뷰어의 산단 단위 "
          "추정 소비량과 1:1 비교 대상이 아님")

CODE_MAP = {
    "당진시": "44270", "서산시": "44210", "아산시": "44200", "천안시": "44130",
    "예산군": "44810", "홍성군": "44800", "보령시": "44180",
    "화성시": "41590", "평택시": "41220", "용인시": "41463",
    "시흥시": "41390", "안산시": "41390",
    "이천시": "41500", "파주시": "41480", "김포시": "41570",
}

BROKEN_FILES = set()  # 전국/오export로 확인된 시군명을 임시로 등록(재다운로드 대기 중)


def to_number(cell):
    if cell is None:
        return 0
    return float(str(cell).replace(",", "").strip())


def parse_file(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    total_kwh = 0.0
    manuf_kwh = 0.0
    for r in rows[4:]:
        category = (r[2] or "").strip()
        if not category:
            continue
        usage = to_number(r[4])
        if category == "제조업":
            manuf_kwh += usage
    total_row = rows[-1]
    total_kwh = to_number(total_row[4])
    return total_kwh, manuf_kwh


def main():
    by_code = {}
    skipped = []
    for path in sorted(glob.glob(os.path.join(USAGE_DIR, "*.xlsx"))):
        fname = os.path.basename(path)
        city = fname.split(" ")[0]
        if city in BROKEN_FILES:
            skipped.append(fname)
            continue
        code = CODE_MAP.get(city)
        if not code:
            print(f"  경고: 미매핑 시군명 '{city}' ({fname}) — 건너뜀")
            continue
        total_kwh, manuf_kwh = parse_file(path)
        entry = by_code.setdefault(code, {"total_kwh_3yr": 0.0, "manuf_kwh_3yr": 0.0, "source_files": []})
        entry["total_kwh_3yr"] += total_kwh
        entry["manuf_kwh_3yr"] += manuf_kwh
        entry["source_files"].append(fname)
        print(f"  {fname:45s} -> code {code}  total={total_kwh:,.0f}kWh(3yr)  manuf={manuf_kwh:,.0f}kWh(3yr)")

    out = {
        "period_label": PERIOD_LABEL,
        "source_label": SOURCE_LABEL,
        "caveat": CAVEAT,
        "by_code": {},
    }
    for code, e in by_code.items():
        out["by_code"][code] = {
            "total_gwh_year": round(e["total_kwh_3yr"] / 3 / 1_000_000, 2),
            "manuf_gwh_year": round(e["manuf_kwh_3yr"] / 3 / 1_000_000, 2),
            "source_files": e["source_files"],
            "note": None,
        }
    for city in BROKEN_FILES:
        code = CODE_MAP.get(city)
        if code and code not in out["by_code"]:
            out["by_code"][code] = {
                "total_gwh_year": None,
                "manuf_gwh_year": None,
                "source_files": [],
                "note": "원본 파일이 손상되어(전국/오export 등) 재다운로드 대기 중",
            }

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"\n완료: {len(out['by_code'])}개 카드 코드 -> {OUT_PATH}")
    if skipped:
        print(f"제외된 파일(손상 의심): {skipped}")


if __name__ == "__main__":
    main()
