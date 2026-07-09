#!/usr/bin/env python3
"""metadata/kepco_usage/*.xlsx(한전 산업분류별 전력사용량, 2023.01~2025.12 누적)에서
시군별 **업종(KSIC 대분류) 전력수요 프로필**을 만든다 → kepco_sector_profile.json
(루트, 커밋 대상 아님 — Supabase 업로드 대상, ROOT_FILES에 등록됨).

build_kepco_usage.py(제조업/전체 합계만 산출)와 원본·매핑 로직을 공유하되,
19개 업종 전체의 연평균 GWh와 비중, 상위 업종을 담는다 — "그 지역의 주요
전력 수요 업종이 무엇인가"를 보여주는 것이 목적(insight.html 시각화용).

시흥+안산(1)(2)=41390, 용인(1)(2)=41463 합산 규칙은 build_kepco_usage.py와 동일.
"""
import glob
import json
import os

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
USAGE_DIR = os.path.join(ROOT, "metadata", "kepco_usage")
OUT_PATH = os.path.join(ROOT, "kepco_sector_profile.json")

YEARS = 3.0  # 2023.01~2025.12 누적 → 연평균 환산
PERIOD_LABEL = "2023.01–2025.12(3개년 누적치의 연평균)"
SOURCE_LABEL = "한국전력공사 빅데이터센터(bigdata.kepco.co.kr), 산업분류별 전력사용량 2023–2025 연평균"
CAVEAT = "시군구 전체 계약(KSIC 대분류) 합산치 — 산단 단위 추정 소비량과 1:1 비교 대상 아님"

CODE_MAP = {
    "당진시": "44270", "서산시": "44210", "아산시": "44200", "천안시": "44130",
    "예산군": "44810", "홍성군": "44800", "보령시": "44180",
    "화성시": "41590", "평택시": "41220", "용인시": "41463",
    "시흥시": "41390", "안산시": "41390",
    "이천시": "41500", "파주시": "41480", "김포시": "41570",
}


def to_number(cell):
    if cell is None:
        return 0
    return float(str(cell).replace(",", "").strip())


def parse_file(path):
    """{업종명: 누적 kWh} 반환. 업종명은 공백 정리만(원문 표기 유지)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    out = {}
    for r in list(ws.iter_rows(values_only=True))[4:]:
        if len(r) < 5:
            continue
        category = (r[2] or "").strip()
        if not category:
            continue
        out[category] = out.get(category, 0.0) + to_number(r[4])
    return out


def main():
    by_code = {}
    for path in sorted(glob.glob(os.path.join(USAGE_DIR, "*.xlsx"))):
        sgg_name = os.path.basename(path).split(" ")[0]
        code = CODE_MAP.get(sgg_name)
        if not code:
            print(f"  [스킵] 매핑 없는 파일: {os.path.basename(path)}")
            continue
        sectors = parse_file(path)
        acc = by_code.setdefault(code, {})
        for cat, kwh in sectors.items():
            acc[cat] = acc.get(cat, 0.0) + kwh

    out_codes = {}
    for code, acc in by_code.items():
        gwh = {cat: kwh / YEARS / 1e6 for cat, kwh in acc.items()}  # kWh누적 → GWh/년
        total = sum(gwh.values())
        ranked = sorted(gwh.items(), key=lambda x: -x[1])
        out_codes[code] = {
            "total_gwh_year": round(total, 1),
            "sectors_gwh_year": {c: round(v, 1) for c, v in ranked},
            "top3": [
                {"sector": c, "gwh_year": round(v, 1),
                 "share_pct": round(v / total * 100, 1) if total else 0}
                for c, v in ranked[:3]
            ],
            "manuf_share_pct": round(gwh.get("제조업", 0) / total * 100, 1) if total else 0,
        }

    out = {
        "period_label": PERIOD_LABEL,
        "source_label": SOURCE_LABEL,
        "caveat": CAVEAT,
        "by_code": out_codes,
    }
    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"완료: {len(out_codes)}개 시군 -> {OUT_PATH}")
    for code, e in sorted(out_codes.items()):
        top = e["top3"][0]
        print(f"  {code}: 총 {e['total_gwh_year']:,.0f} GWh/년, 1위 {top['sector']} {top['share_pct']}% (제조업 {e['manuf_share_pct']}%)")


if __name__ == "__main__":
    main()
